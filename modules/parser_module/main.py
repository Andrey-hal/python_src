print("Проверка наличия нужных библиотек.")
try:
    import os.path
    import sys

    from openpyxl import load_workbook, Workbook
    from json import dumps
    from subprocess import Popen, PIPE
    from definitions import *
    from logs import logging, bclr
    from models.work_db import *
    from models.local_db import *
    from modules.parser_module.utils import string_tools
    from modules.parser_module.utils import parsing_rules
    from modules.parser_module.utils import table_tools
    from modules.parser_module.utils.xlsx_tools import pars_formated_strings
    from modules.parser_module.utils import other_tools
except:
    from traceback import format_exc

    print("Ошибка запуска. Одна или несколько библиотек отсутствуют. {}".format(format_exc()))
    raise SystemExit()

log = logging.getLogger("parser")
log.propagate = False
config = Conf("config.ini").config


class MainParser:
    """
    Класс парсинга переменных из файла таблиц типа .xlsx.
    """

    def __init__(self, file=CONFIG_DIR+config['parsing']['input_file_name']):

        if os.path.isfile(file):
            log.info(
                "Проверка наличия файла первоначальной конфигурации... в папке {}\n".format(CONFIG_DIR)
                + (bclr.OKG + "OK" + bclr.C)
            )
        else:
            log.critical(
                "Проверка наличия файла первоначальной конфигурации... в папке {}\n".format(CONFIG_DIR)
                + (bclr.F + "Ошибка" + bclr.C)
            )
            raise SystemExit

        workbook = load_workbook(filename=file, data_only=True)  # Книга для данных в ячейках
        self.sheet = workbook.worksheets[0]

        workbook_formls = load_workbook(filename=file)  # Книга для формул в них
        self.sheet_formls = workbook_formls.worksheets[0]

        self.sheet_arrays_formls = {  # Массив с формулами-массивами
            k: v
            for k, v in self.sheet_formls.formula_attributes.items()
            if v["t"] == "array"
        }

        self.tables = {}
        self.variables = {}
        self.errors = []

        self.success = True
        self.latex_formated_strings = pars_formated_strings(file)

    def make_error(self, text, error_type, n, traceback=None):
        self.errors.append({"coordinate": n, "text": text, "type": error_type})
        if error_type == "critical":
            self.success = False
        log.error("#{}: ".format(n) + text)
        log.debug(traceback)

    def table_parsing(self):
        result = table_tools.table_parsing(self.sheet_formls)
        if len(result["errors"]):
            for err in result["errors"]:
                self.make_error(**err)
        if not len(result["tables"]):
            return False
        self.tables = result["tables"]
        return True

    def vars_parsing(self):
        result = table_tools.vars_parsing(self)
        self.variables.update(
            table_tools.array_expressions_parsing(
                self.sheet_arrays_formls, self.sheet_formls
            )
        )

    def fill_latex_equations(self):
        """
        Проходимся по переменным и присваиваем им LaTex формулу
        """
        for row_n, value_data in self.variables.items():
            try:
                formls = string_tools.expression_decoder_latex_equation(
                    value_data["formulae"], self.variables
                )
            except KeyError:
                if "internal_var" in value_data["name"]:
                    continue
                pass
            except:
                log.error('Ошибка парсинга Latex формулы в строке {}'.format(row_n))
                continue

            if formls is not "":
                self.variables[row_n]["latex_equation"] = formls
            else:
                self.variables[row_n]["latex_equation"] = value_data["latex_symbol"]

    def upload_to_php(self):
        result = other_tools.upload_to_php(self.tables)
        if result:
            return True
        else:
            self.make_error(**result[1])
            return False

    @db_session
    def upload_to_base(self):
        old_local_vars = [v_l.name for v_l in select(f for f in Variable)]
        old_local_vars.extend([v['name'] for k,v in self.variables.items()])
        delete(v for v in VariableWork if v.name in old_local_vars)
        commit()

        tables = self.tables
        for table_n, (table_name, table) in enumerate(tables.items(), start=1):
            for var_row, var in table['variables'].items():
                # VariableWork.get(name=var['name']).delete()
                try:
                    VariableWork(
                        name=var['name'],
                        description=var['description'],
                        sign=var['symbol'],
                        units=var['units'],
                        type=var['data_type'],
                        is_array=True if len(var['value']) > 1 or len(var['expression']) > 1 else False,
                        flo=[i for i in var['value'].values() if i is not None],
                        need_to_logbox = var['need_to_logbox']
                    )
                except pony.orm.core.CacheIndexError:
                    log.error('Ошибка уникальности для переменной {}'.format(var['name']))
        VariableWork(name=None)
        commit()

    @db_session
    def generate_local_db(self):
        """
        А тут делаем хорошее дело: пишем всё в локальную базу, чтобы хоть где-то это хранилось структурированно.
        """
        delete(t for t in Table)
        delete(t for t in Variable)
        commit()
        # таблицы
        for k, v in self.tables.items():
            v["name"] = k

            del v["variables"]
            Table(**v)
        commit()

        # переменные
        for k_v, v_v in self.variables.items():

            try:
                table = Table.get(name=v_v["table_name"])
                # v_v["name"] = "Tabl{}_{}".format(
                #     list(self.tables.keys()).index(v_v["table_name"]) + 1, v_v["name"]
                # )
            except:
                table = None

            list(map(lambda x: v_v.pop(x, None),["table_name", "formulae", "latex_equation"],))  # Убираем то, что не нужно писать в базу.
            Variable(
                **v_v, table=table, static=False if (len(v_v["expression"]) or len(v_v["condition"])) else True
            )

        commit()

    def start(self):  # TODO Нормально назвать функцию
        """
        Поочередно запускаем все функции для первоначальной записи переменных в базу.
        """

        log.info("Запуск парсера. Версия")
        log.info("Запуск парсинга таблиц...")

        if self.table_parsing():
            log.info(bclr.OKG + "\rOK" + bclr.C)
            log.info("Таблиц успешно спарсено: {}".format(len(self.tables)))
        else:
            log.critical(
                bclr.F
                + "Ошибка в парсинге таблиц. Удалось спарсить {} таблиц".format(
                    len(self.tables)
                )
                + bclr.C
            )

        log.info("Запуск парсинга переменных...")
        self.vars_parsing()

        if self.success:
            log.info(bclr.OKG + "\rOK" + bclr.C)
            log.info("Переменных успешно спарсено: {}".format(len(self.variables)))
        else:
            log.critical(
                bclr.F + "Ошибка в парсинге переменных. Удалось спарсить {} "
                         "переменных".format(len(self.variables)) + bclr.C
            )

        log.info("Запуск парсинга LaTex формул")
        self.fill_latex_equations()
        log.info("Запись переменных в php...")

        self.upload_to_php()
        log.info("Запись переменных в базу...")
        self.upload_to_base()
        self.generate_local_db()
        log.info('Записано')

if __name__ == "__main__":
    m = MainParser()
    m.start()
